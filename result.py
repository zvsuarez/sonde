# -*- coding: utf-8 -*-

# Form implementation generated from reading ui file 'ui/result.ui'


from PyQt5 import QtCore, QtGui, QtWidgets
from openpyxl import Workbook

class TableView(QtWidgets.QTableWidget):

    def setData(self):
        headers = []
        for n, key in enumerate(self.data.keys()):
            headers.append(key)
            for m, item in enumerate(self.data[key]):
                newItem = QtWidgets.QTableWidgetItem(item)
                self.setItem(m, n, newItem)
        self.setHorizontalHeaderLabels(headers)

    def setDimensions_col(self):
        total_col = 0
        for column in range(QtWidgets.QTableWidget.columnCount(self)):
            total_col += QtWidgets.QTableView.columnWidth(self,column)

        if total_col > 1280:
            return 1280
        if total_col % 2 != 0:
            return (total_col - 1) + 40
        else:
            return total_col + 40

    def setDimensions_row(self):
        total_row = 0
        for row in range(QtWidgets.QTableWidget.rowCount(self)):
            total_row += QtWidgets.QTableView.rowHeight(self,row)

        if total_row > 720:
            return 720
        if total_row % 2 != 0:
            return (total_row - 1) + 25
        else:
            return total_row + 25
        
    def __init__(self, data, *args):
        QtWidgets.QTableWidget.__init__(self, *args)
        self.data = data
        self.setData()
        self.resizeColumnsToContents()
        self.resizeRowsToContents()
        col = self.setDimensions_col()
        row = self.setDimensions_row()
        self.setGeometry(QtCore.QRect(0, 0, col, row))
        #print(col, row)


class Ui_Dialog(object):

    def save_result(self):
        dialog = QtWidgets.QFileDialog()
        path = dialog.getSaveFileName(QtWidgets.QMainWindow(), 'Save File...', '', filter='Excel File (*.xlsx *.xls)')
        wb = Workbook()
        ws = wb.active

        try:
            with open(path[0], 'wb') as f:
                for column in range(self.tableWidget.columnCount()):
                    ws.cell(1, column+1, self.tableWidget.horizontalHeaderItem(column).text())

                for row in range(self.tableWidget.rowCount()):
                    for col in range(self.tableWidget.columnCount()):
                        item = self.tableWidget.item(row, col)
                        if item is not None:
                            ws.cell(row+2, col+1, item.text())
                        else:
                            ws.cell(row+2, col+1, '')
                wb.save(f)

        except FileNotFoundError:
            pass

    def setupUi(self, Dialog, data, *args):
        self.tableWidget = TableView(data, *args, Dialog)
        self.tableWidget.setObjectName("tableWidget")
        #self.tableWidget = QtWidgets.QTableWidget(Dialog)
        #self.tableWidget.setGeometry(QtCore.QRect(0, 0, 1280, 660))
        #self.tableWidget.setColumnCount(0)
        #self.tableWidget.setRowCount(0)
        Dialog.setObjectName("Dialog")
        Dialog.setFixedSize(self.tableWidget.setDimensions_col(), self.tableWidget.setDimensions_row()+80)
        Dialog.setWindowFlag(QtCore.Qt.WindowContextHelpButtonHint, False)
        icon = QtGui.QIcon()
        icon.addPixmap(QtGui.QPixmap(":/icons/ui/icons/32x32.png"), QtGui.QIcon.Normal, QtGui.QIcon.Off)
        icon.addPixmap(QtGui.QPixmap(":/icons/ui/icons/32x32.png"), QtGui.QIcon.Normal, QtGui.QIcon.On)
        Dialog.setWindowIcon(icon)
        self.buttonBox = QtWidgets.QDialogButtonBox(Dialog)
        self.buttonBox.setGeometry(QtCore.QRect((self.tableWidget.setDimensions_col() // 2)-100, self.tableWidget.setDimensions_row() + ((self.tableWidget.setDimensions_row()+60 - self.tableWidget.setDimensions_row()) / 2) - 10, 200, 40))
        font = QtGui.QFont()
        font.setFamily("MS Shell Dlg 2")
        self.buttonBox.setFont(font)
        self.buttonBox.setOrientation(QtCore.Qt.Horizontal)
        self.buttonBox.setStandardButtons(QtWidgets.QDialogButtonBox.Cancel|QtWidgets.QDialogButtonBox.Save)
        self.buttonBox.setObjectName("buttonBox")
        
        self.retranslateUi(Dialog)
        self.buttonBox.accepted.connect(self.save_result)
        self.buttonBox.rejected.connect(Dialog.reject)
        QtCore.QMetaObject.connectSlotsByName(Dialog)

    def retranslateUi(self, Dialog):
        _translate = QtCore.QCoreApplication.translate
        Dialog.setWindowTitle(_translate("Dialog", "Decode Result"))
import resource


if __name__ == "__main__":
    import sys
    app = QtWidgets.QApplication(sys.argv)
    Dialog = QtWidgets.QDialog()
    ui = Ui_Dialog()
    ui.setupUi(Dialog, data=dict())
    Dialog.show()
    sys.exit(app.exec_())
