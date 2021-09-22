# -*- coding: utf-8 -*-

# Form implementation generated from reading ui file 'ui/decode_window.ui'


from PyQt5 import QtCore, QtGui, QtWidgets
from openpyxl import Workbook
import os

class TableView(QtWidgets.QTableWidget):
    def __init__(self, data, *args):
        QtWidgets.QTableWidget.__init__(self, *args)
        self.setGeometry(QtCore.QRect(0, 0, 1060, 760))
        self.data = data
        self.setData()
        self.resizeColumnsToContents()
        self.resizeRowsToContents()
    def setData(self):
        headers = []
        for n, key in enumerate(self.data.keys()):
            headers.append(key)
            for m, item in enumerate(self.data[key]):
                newItem = QtWidgets.QTableWidgetItem(item)
                self.setItem(m, n, newItem)
        self.setHorizontalHeaderLabels(headers)

class Ui_Dialog(object):

    def save_result(self):
        dialog = QtWidgets.QFileDialog()
        path = dialog.getSaveFileName(QtWidgets.QMainWindow(), 'Save File...', '', filter='Excel Files (*xlsx *xls)')
        wb = Workbook()
        ws = wb.active

        try:
            with open(path[0], 'wb') as f:
                for column in range(self.tableWidget.columnCount()):
                    ws.cell(1, column+1, self.tableWidget.horizontalHeaderItem(column).text())

                for row in range(self.tableWidget.rowCount()):
                    for col in range(self.tableWidget.columnCount()):
                        item = self.tableWidget.item(row, col)
                        if item is not None:
                            ws.cell(row+2, col+1, item.text())
                        else:
                            ws.cell(row+2, col+1, '')
                wb.save(f)

        except FileNotFoundError:
            pass
        
    def setupUi(self, Dialog, data, *args):
        Dialog.setObjectName("Dialog")
        Dialog.setFixedSize(1060, 820)
        Dialog.setWindowFlag(QtCore.Qt.WindowContextHelpButtonHint, False)
        icon = QtGui.QIcon()
        icon.addPixmap(QtGui.QPixmap(":/icons/32x32.png"), QtGui.QIcon.Normal, QtGui.QIcon.Off)
        Dialog.setWindowIcon(icon)
        self.buttonBox = QtWidgets.QDialogButtonBox(Dialog)
        self.buttonBox.setGeometry(QtCore.QRect(440, 770, 200, 41))
        font = QtGui.QFont()
        font.setFamily("MS Shell Dlg 2")
        self.buttonBox.setFont(font)
        self.buttonBox.setOrientation(QtCore.Qt.Horizontal)
        self.buttonBox.setStandardButtons(QtWidgets.QDialogButtonBox.Cancel|QtWidgets.QDialogButtonBox.Save)
        self.buttonBox.setObjectName("buttonBox")
        self.tableWidget = TableView(data, *args, Dialog)
        #self.tableWidget = QtWidgets.QTableWidget(Dialog)
        #self.tableWidget.setGeometry(QtCore.QRect(0, 0, 1060, 760))
        self.tableWidget.setObjectName("tableWidget")
        #self.tableWidget.setColumnCount(0)
        #self.tableWidget.setRowCount(0)

        self.retranslateUi(Dialog)
        self.buttonBox.accepted.connect(self.save_result)
        self.buttonBox.rejected.connect(Dialog.reject)
        QtCore.QMetaObject.connectSlotsByName(Dialog)

    def retranslateUi(self, Dialog):
        _translate = QtCore.QCoreApplication.translate
        Dialog.setWindowTitle(_translate("Dialog", "Decode Result"))
import resource


if __name__ == "__main__":
    import sys
    app = QtWidgets.QApplication(sys.argv)
    Dialog = QtWidgets.QDialog()
    ui = Ui_Dialog()
    ui.setupUi(Dialog, data=dict())
    Dialog.show()
    sys.exit(app.exec_())
